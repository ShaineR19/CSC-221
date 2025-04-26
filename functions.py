# -*- coding: utf-8 -*-
"""
Takes data from deanDailyCsar and FTE_Tier to determine classes FTE.
Allows the user to Get Course Enrollment, and get FTE by Division, Instructor,
and Course.

GROUP A
Teresa Hearn, Shiane Ransford, Latoya Winston

3/18/2025

CSC-221-001

M4GroupAPro

"""
import traceback
import pandas as pd
from openpyxl import load_workbook

def menu():
    '''
    Displays the menu options.

    Returns
    -------
    None.F

    '''
    print()
    print("="*20+"Menu"+'='*20)
    print('1) Enter "Sec Divisions" code ')
    print('2) Get course Enrollment Percentage')
    print('3) FTE by Division')
    print('4) FTE per instructor (for specific Div)')
    print('5) FTE per course (for specific Div)')
    print('6) Exit')
    print("="*44)

def readfile():
    '''
    Generates the dataframe and then sorts it.

    Returns
    -------
    groups : dataframe
        the sorted dataframe of the file.

    '''
    try:
        # reads the deansDailyCsar.csv and unique_deansDailyCsar_FTE files in to a dataframe
        file_in = pd.read_csv('deanDailyCsar.csv')
        fte_file_in = pd.read_excel('unique_deansDailyCsar_FTE.xlsx')

        # merge prior dataframes
        # Extract Course Code from Sec Name if not already done
        if "Course Code" not in file_in.columns:
            file_in["Course Code"] = file_in["Sec Name"].str.extract(r"([A-Z]{3}-\d{3})")

        # Also create Course Code in credits_df
        if "Course Code" not in fte_file_in.columns:
            fte_file_in["Course Code"] = fte_file_in["Sec Name"].str.extract(r"([A-Z]{3}-\d{3})")

        # Merge only needed columns from credits_df
        merged_df = pd.merge(
            file_in,
            fte_file_in[["Course Code", "Contact Hours"]],
            how='left',
            on='Course Code'
        )

        merged_df["Contact Hours"] = pd.to_numeric(
        merged_df["Contact Hours"], errors='coerce')
        merged_df["FTE Count"] = pd.to_numeric(merged_df["FTE Count"], errors='coerce')

        # Calculate Total FTE
        merged_df["Total FTE"] = ((merged_df["Contact Hours"] * 16 * merged_df["FTE Count"]) / 512).round(3)


        # sorts the dataframe by sec divisions, sec name
        # and sec faculty info and assigns it to groups
        groups = file_in.sort_values(["Sec Divisions", "Sec Name",
                                      "Sec Faculty Info"])

        return groups

    except FileNotFoundError:
        groups = []
        print("File Missing!")
        return groups

def SecDivisions(file_in):
    '''
    Allows user to enter sec divisions to search for.

    Parameters
    ----------
    file_in : dataframe
        Contains the information for each division.

    Returns
    -------
    None.

    '''
    try:
        # access and displays the available sec divisons
        # to choose from in rows of 4.
        print()
        print("Available Sec Divisions: \n")
        sec_group = sorted(file_in["Sec Divisions"].dropna().unique())

        # Display divisions in rows of 4
        for i in range(0, len(sec_group), 4):
            row_items = sec_group[i:i+4]
            for x in row_items:
                print(f'-{x}', end=' ')
            print()

        # Get and process users input
        sec_input = input("\nEnter Sec Divisions separated by commas or ALL: ")
        sec_input = sec_input.upper().strip()

        if sec_input == 'ALL':  # Check for ALL before splitting
            divisions_to_process = sec_group
        else:
            divisions_to_process = sec_input.split(',')  # Split only if not ALL
            divisions_to_process = [div.strip() for div in divisions_to_process]

        # Validate and process divisions
        for div in divisions_to_process:
            if div.upper() not in [d.upper() for d in sec_group]:
                print(f"\nWarning: Division '{div}' not found")

        
        for division in divisions_to_process:
            if division.upper() in [d.upper() for d in sec_group]:
                # actual_division = next(d for d in sec_group if d.upper() == division.upper())
                # process_division(file_in, actual_division)

                # Extract rows for selected division into a new dataframe
                # Convert division name to lowercase for dataframe name
                df_names = division.lower()
                df_name = file_in[file_in['Sec Divisions'] == division].copy()

                # Create Excel filename (lowercase)
                excel_filename = f"{division.lower()}.xlsx"

                # Write to Excel
                df_name.to_excel(excel_filename, index=False)

                print(f"\nCreated DataFrame '{df_names}' with {len(df_name)} rows")
                print(f"Saved to file: {excel_filename}")

    except TypeError:
        print("Missing information from file. Check to be sure the file is not missing.")
    except Exception as err:
        print("Error: "+str(err))

def option2_enrollment(df):
    '''
    Parameters
    ----------
    df : dataframe
        file data fram deansDailyCsar.csv

    Returns
    -------
    Returns course enrollment percentage

    '''
    course_code = True

    while course_code:
        course_input = input("Enter course code (e.g., ACA-120) " \
            "or type 'back' to return: ").strip()

        if course_input.lower() == 'back':
            return

        # Filter rows in 'Sec Name' containing the course code (case insensitive)
        filtered_df = df[df["Sec Name"].str.contains(course_input, case=False, na=False)]

        if filtered_df.empty:
            print("Course not found. Please re-enter the course code or"
                  "type 'back' to return to the main menu.")
        else:
            course_code = False

    # For face-to-face sections, duplicate rows may exist.
    # Here, we drop duplicate rows based on 'Sec Name'.
    # This assumes online sections (which contain a '9' in the section number)
    # are unique or do not duplicate.
    filtered_df = filtered_df.drop_duplicates(subset="Sec Name")

    # Define a function to calculate enrollment percentage for a row
    def calc_enrollment(row):
        try:
            cap = float(row["Capacity"])
            fte = float(row["FTE Count"])

            if cap == 0:
                return "0%"

            percentage = (fte / cap) * 100
            return f"{percentage:.2f}%"

        except (ValueError, TypeError, ZeroDivisionError):
            return "N/A%"

    # Calculate and add the Enrollment Percentage column
    filtered_df["Enrollment Percentage"] = filtered_df.apply(calc_enrollment, axis=1)

    # Create the output DataFrame with the required columns
    output_columns=[
        "Sec Name",
        "X Sec Delivery Method",
        "Meeting Times",
        "Capacity",
        "FTE Count",
        "Total FTE",
        "Sec Faculty Info",
        "Enrollment Percentage"
    ]
    output_df = filtered_df[output_columns]

    # Determine the file name based on the course code entered
    file_code = course_input.replace("-", "").lower()
    file_name = f"{file_code}_per.xlsx"

    # Put output.df into an excel file
    output_df.to_excel(file_name, index=False)

    # Adjust column widths with openpyxl workbook
    wb = load_workbook(file_name)
    ws = wb.active

    for column_cells in ws.columns:
        column_letter = column_cells[0].column_letter
        ws.column_dimensions[column_letter].width = 25

    # Save the File
    wb.save(file_name)
    print(f"Created '{file_name}' with enrollment data.")

def division_fte(file_in):
    '''
    Analyze FTE by Division and export to a sheet in a division-specific report file.

    Parameters
    ----------
    file_in : pandas.DataFrame
        Input DataFrame containing course information.

    Returns
    -------
    None
    '''
    print()
    # Get unique division codes
    divisions = sorted(file_in['Sec Divisions'].dropna().unique())

    # Display available divisions
    print("Available Division Codes:")
    for i in range(0, len(divisions), 4):
        row = divisions[i:i+4]
        print("  ".join(f"{div}" for div in row))

    # Get division code from user
    div_code = input("\nEnter Division Code: ").strip()

    if not div_code:
        print("Please enter a valid division code.")
        return

    # Convert to uppercase for case-insensitive comparison
    div_code = div_code.upper()

    # Check if division exists (case-insensitive comparison)
    valid_divisions = [div.upper() for div in divisions]
    if div_code not in valid_divisions:
        print(f"Division '{div_code}' not found. Please check the code and try again.")
        return

    # Read FTE tier data
    fte_data = pd.read_excel('FTE_Tier.xlsx')

    # Create a lookup dictionary for faster access
    fte_lookup = {row['Prefix/Course ID']: row['New Sector']
                 for _, row in fte_data.iterrows() if not pd.isna(row['Prefix/Course ID'])}

    # Get the actual division code with correct case
    actual_div = divisions[valid_divisions.index(div_code)]

    try:
        # Filter data for the selected division
        div_data = file_in[file_in['Sec Divisions'] == actual_div].copy()

        if len(div_data) == 0:
            print("No data found for this division.")
            return

        # Add course code column
        div_data['Course Code'] = div_data['Sec Name'].str.extract(r'([A-Z]+-\d+)')

        # Sort by Course Code and Sec Name
        div_data = div_data.sort_values(['Course Code', 'Sec Name'])

        # Base value for FTE calculation
        base_fte_value = 1926

        # Create output list to store rows
        output_rows = []
        current_course = None
        course_total_fte = 0
        first_row = True

        # Process each row
        for _, row in div_data.iterrows():
            course = row['Course Code']

            # If new course and not first course, add total for previous course
            if course != current_course and current_course is not None:
                output_rows.append({
                    'Division': '',
                    'Course Code': 'Total',
                    'Sec Name': '',
                    'X Sec Delivery Method': '',
                    'Meeting Times': '',
                    'Capacity': '',
                    'FTE Count': '',
                    'Sec Faculty Info': '',
                    'Total FTE': '',
                    'Enrollment Per': '',
                    'Generated FTE': course_total_fte
                })
                course_total_fte = 0

            # Get section prefix (first 3 characters of section name)
            sec = row['Sec Name'][:3] if not pd.isna(row['Sec Name']) else ""

            # Look up new sector value from the dictionary
            new_sector_value = fte_lookup.get(sec, 0)

            # Calculate adjusted FTE for the current row
            total_fte = float(row['Total FTE'])
            
            #Calculate adjusted_fte
            adjusted_fte = (total_fte * (new_sector_value + base_fte_value))

            # Calculate enrollment percentage
            enrollment_per = ''
            if pd.notna(row['Capacity']) and pd.notna(row['FTE Count']) and \
                        float(row['Capacity']) > 0:

                enrollment_per = (float(row['FTE Count']) / float(row['Capacity'])) * 100
                enrollment_per = f"{round(enrollment_per, 2)}%"  # Add percentage sign here

            # Add current row with enrollment percentage and generated FTE
            output_rows.append({
                'Division': actual_div if first_row else '',
                'Course Code': course if course != current_course else '',
                'Sec Name': row['Sec Name'],
                'X Sec Delivery Method': row['X Sec Delivery Method'],
                'Meeting Times': row['Meeting Times'],
                'Capacity': row['Capacity'],
                'FTE Count': row['FTE Count'],
                'Sec Faculty Info': row['Sec Faculty Info'],
                'Total FTE': row['Total FTE'],
                'Enrollment Per': enrollment_per,
                'Generated FTE': adjusted_fte
            })

            # Add to course total
            course_total_fte += adjusted_fte

            current_course = course
            first_row = False

        # Add total for last course
        if current_course is not None:
            output_rows.append({
                'Division': '',
                'Course Code': 'Total',
                'Sec Name': '',
                'X Sec Delivery Method': '',
                'Meeting Times': '',
                'Capacity': '',
                'FTE Count': '',
                'Sec Faculty Info': '',
                'Total FTE': '',
                'Enrollment Per': '',
                'Generated FTE': course_total_fte
            })

        # Format Generated FTE column with '$' before the number
        # and a comma every 3 digits
        for row in output_rows:
            row['Generated FTE'] = f"${row['Generated FTE']:,.2f}"

        # Convert to DataFrame
        output_df = pd.DataFrame(output_rows)

        # Create Excel file
        excel_filename = f"{actual_div.lower()}_fte.xlsx"

        # Write to Excel
        with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
            output_df.to_excel(writer, sheet_name='Division Analysis', index=False)

            # Format the worksheet
            worksheet = writer.sheets['Division Analysis']

            # Adjust column widths
            for idx, col in enumerate(output_df.columns):
                max_length = max(
                    output_df[col].astype(str).apply(len).max(),
                    len(str(col))
                ) + 2
                worksheet.column_dimensions[chr(65 + idx)].width = max_length

        print(f"\nAnalysis for division: {actual_div}")
        print(f"Results exported to {excel_filename}")

    except (FileNotFoundError, ValueError, KeyError):
        print("Error processing data")
        print(traceback.format_exc())
        return

    return


def clean_name_for_search(name):
    '''
    Standardize name format for searching.
    Removes periods and extra spaces.

    Parameters
    ----------
    name : str
        Name to clean

    Returns
    -------
    str
        Cleaned name for comparison
    '''
    return name.replace('.', '').strip().lower()


def clean_instructor_name(name):
    '''
    Clean instructor name for file naming.
    Handles different formats like "H Seidi", "H. Seidi", etc.

    Parameters
    ----------
    name : str
        Instructor name

    Returns
    -------
    str
        Cleaned name formatted for filename
    '''
    # Split by comma first if it exists
    if ',' in name:
        last_name, first_part = name.split(',', 1)
        # Clean up the last name and first initial
        last_name = last_name.strip().lower()
        # Get first character and remove any periods
        first_initial = first_part.strip().replace('.', '')[0].lower()
    else:
        # Handle space-separated names
        parts = name.split()
        last_name = parts[-1].lower()
        # Get first character and remove any periods
        first_initial = parts[0].replace('.', '')[0].lower()

    return f"{last_name}{first_initial}_FTE.xlsx"


def instructorFTE(file_in):
    '''
    Analyze FTE by instructor within a division.

    Parameters
    ----------
    file_in : pandas.DataFrame
        Input DataFrame containing course information.

    Returns
    -------
    None
    '''
    print()
    # Get unique faculty names for reference
    faculty = sorted(file_in['Sec Faculty Info'].dropna().unique())
    print(f"Found {len(faculty)} faculty members")

    while True:
        print("\nEnter instructor name (first or last name)")
        print("Type 'list' to see all instructors")
        print("Type 'back' for main menu")

        faculty_name = input("\nEnter name: ").strip()

        if faculty_name.lower() == 'back':
            print("\nReturning to main menu...")
            return

        if faculty_name.lower() == 'list':
            print("\nInstructors:")
            for i in range(0, len(faculty), 3):
                names = faculty[i:i+3]
                print("  ".join(f"{name:<30}" for name in names))
            continue

        if not faculty_name:
            print("Please enter a valid name.")
            continue

        # Case-insensitive search for partial matches with standardized format
        cleaned_input = clean_name_for_search(faculty_name)
        matches = [f for f in faculty if cleaned_input
                   in clean_name_for_search(f)]
        print(f"Found {len(matches)} matching instructors")

        if not matches:
            print(f"No instructors found matching '{faculty_name}'.")
            print("Try searching without periods (.) or check \
the instructor list.")
            continue

        if len(matches) > 1:
            print("\nMultiple instructors found:")
            for i, name in enumerate(matches, 1):
                print(f"{i}. {name}")
            choice = input("\nEnter number to select \
instructor (or press Enter to search again): ")
            if not choice.isdigit() or int(choice) < 1\
                    or int(choice) > len(matches):
                continue
            selected_faculty = matches[int(choice) - 1]
        else:
            selected_faculty = matches[0]

        print(f"\nProcessing data for: {selected_faculty}")

        try:
            # Filter data for selected instructor
            instructor_data = file_in[file_in['Sec Faculty Info']
                                      == selected_faculty].copy()
            print(f"Found {len(instructor_data)} initial rows")

            if len(instructor_data) == 0:
                print("No courses found for this instructor.")
                continue

            # Add course code column
            instructor_data['Course Code'] = \
                instructor_data['Sec Name'].str.extract(r'([A-Z]+-\d+)')

            # Sort by Course Code and Sec Name
            instructor_data = \
                instructor_data.sort_values(['Course Code', 'Sec Name'])

            # Prepare data for Excel
            data_for_excel = []
            current_course = None
            course_total_fte = 0

            for _, row in instructor_data.iterrows():
                course = row['Course Code']

                # If new course and not first course,
                #  add total for previous course
                if course != current_course and current_course is not None:
                    data_for_excel.append([
                        '',                     # Instructor
                        'Total',                # Course Code
                        '',                     # Sec Name
                        '',                     # Delivery Method
                        '',                     # Meeting Times
                        '',                     # Capacity
                        '',                     # FTE Count
                        course_total_fte,       # Total FTE
                        ''                      # Divisions
                    ])
                    course_total_fte = 0

                # Add current row
                data_for_excel.append([
                    selected_faculty if current_course is None else '',
                    course if course != current_course else '',
                    row['Sec Name'],
                    row['X Sec Delivery Method'],
                    row['Meeting Times'],
                    row['Capacity'],
                    row['FTE Count'],
                    row['Total FTE'],
                    row['Sec Divisions']
                ])

                # Update tracking variables
                course_total_fte += float(row['Total FTE'])
                current_course = course

            # Add total for last course
            if current_course is not None:
                data_for_excel.append([
                    '',                     # Instructor
                    'Total',                # Course Code (moved Total here)
                    '',                     # Sec Name
                    '',                     # Delivery Method
                    '',                     # Meeting Times
                    '',                     # Capacity
                    '',                     # FTE Count
                    course_total_fte,       # Total FTE
                    ''                      # Divisions
                ])

            # Convert to DataFrame
            columns = ['Instructor', 'Course Code', 'Sec Name',
                       'X Sec Delivery Method',
                       'Meeting Times', 'Capacity', 'FTE Count',
                       'Total FTE', 'Sec Divisions']
            output_df = pd.DataFrame(data_for_excel, columns=columns)

            # Create Excel file name using the clean_instructor_name function
            excel_filename = clean_instructor_name(selected_faculty)
            print(f"\nCreating file: {excel_filename}")

            # Write to Excel
            output_df.to_excel(excel_filename, index=False)
            print(f"Written {len(output_df)} rows to Excel")

            print(f"\nAnalysis for instructor: {selected_faculty}")
            print(f"Results exported to {excel_filename}")
            break

        except Exception as e:
            print(f"Error processing data: {str(e)}")
            print(traceback.format_exc())
            continue

    return


def clean_course_code(code):

    '''
    Clean course code for file naming.
    Removes dashes and standardizes format.
    '''
    # Remove dash and convert to lowercase
    clean_code = code.replace('-', '').lower()
    return f"{clean_code}_FTE.xlsx"


def fte_per_course(file_in):
    """
    Calculate and export FTE data for a specific course.


    """
    print()
    # Step 1: Extract course codes from section names using regex
    file_in['Course Code'] = file_in['Sec Name'].str.extract(r'([A-Z]+-\d+)')
    course_codes = sorted(file_in['Course Code'].dropna().unique())

    while True:
        # Step 2a: Prompt user for course code input
        print("\nEnter course code (e.g., CSC-121) or\
type 'back' to return to main menu:")
        course_input = input("Course code: ").strip().upper()

        # Step 2b: Allow user to exit function
        if course_input.lower() == 'back':
            return

        # Step 2c: Find matching courses based on partial or complete input
        matching_courses = [c for c in course_codes if course_input in c]

        # Step 2d: Handle case when no matching courses found
        if not matching_courses:
            print(f"No course found with code '{course_input}'.")
            continue

        # Step 2e: Select specific course (or have user select if multiple matches)
        selected_course = matching_courses[0] \
            if len(matching_courses) == 1 else\
            matching_courses[int(input("Enter number to select course: ")) - 1]

        print(f"\nProcessing data for course: {selected_course}")

        try:
            # Step 3a: Filter data for selected course
            course_data = file_in[file_in['Course Code']
                                 == selected_course].copy()
            if course_data.empty:
                print("No sections found for this course.")
                continue

            # Step 3b: Remove duplicate sections
            course_data = course_data.drop_duplicates(subset='Sec Name')

            # Step 3c: Load FTE tier data for calculations
            fte_data = pd.read_excel('FTE_Tier.xlsx')
            fte_lookup = {row['Prefix/Course ID']: row['New Sector']
                          for _, row in fte_data.iterrows()
                          if pd.notna(row['Prefix/Course ID'])}

            # Step 3d: Sort data by section name
            course_data = course_data.sort_values('Sec Name')
            base_fte_value = 1926

            # Step 4a: Initialize variables for calculating FTE
            output_rows = []
            total_generated_fte = 0

            # Step 4b: Process each section to calculate FTE metrics
            for _, row in course_data.iterrows():
                # Extract section prefix (first 3 chars of section name)
                sec_prefix = row['Sec Name'][:3] if not\
                    pd.isna(row['Sec Name']) else ""

                # Look up new sector value for this prefix
                new_sector_value = fte_lookup.get(sec_prefix, 0)

                # Calculate total FTE for this section
                total_fte = float(row['Total FTE']) \
                    if pd.notna(row['Total FTE']) else 0

                # Calculate generated FTE using formula
                generated_fte = total_fte * (new_sector_value + base_fte_value)

                # Calculate enrollment percentage
                enrollment_per = ""
                if pd.notna(row['Capacity']) and\
                        pd.notna(row['FTE Count']) and\
                        float(row['Capacity']) > 0:
                    enrollment_per = round((float(row['FTE Count']) /
                                            float(row['Capacity'])) * 100, 2)

                # Add section data to output rows
                output_rows.append({
                    'Course Code': selected_course
                    if len(output_rows) == 0 else '',
                    'Sec Name': row['Sec Name'],
                    'X Sec Delivery Method': row['X Sec Delivery Method'],
                    'Sec Faculty Info': row['Sec Faculty Info'],
                    'Meeting Times': row['Meeting Times'],
                    'Capacity': row['Capacity'],
                    'FTE Count': row['FTE Count'],
                    'Total FTE': total_fte,
                    'Enrollment Per': f"{enrollment_per}%",
                    'Generated FTE': generated_fte
                })

                # Add current section's FTE to course total
                total_generated_fte += generated_fte

            # Step 4c: Add a summary row with course totals
            output_rows.append({
                'Course Code': 'Total',
                'Sec Name': '',
                'X Sec Delivery Method': '',
                'Sec Faculty Info': '',
                'Meeting Times': '',
                'Capacity': '',
                'FTE Count': '',
                'Total FTE': '',
                'Enrollment Per': '',
                'Generated FTE': total_generated_fte
            })

            for row in output_rows:
                row['Generated FTE'] = f"${row['Generated FTE']:,.2f}"

            # Step 5a: Convert to DataFrame for export
            output_df = pd.DataFrame(output_rows)

            # Step 5b: Create Excel file with course code as name
            file_name = f"{selected_course.replace('-', '').lower()}_FTE.xlsx"
            output_df.to_excel(file_name, index=False)

            # Step 5c: Format Excel file for readability
            wb = load_workbook(file_name)
            ws = wb.active

            for idx, col in enumerate(output_df.columns):
                max_length = max(output_df[col].astype(str).apply(len).max(),
                                 len(str(col))) + 2
                ws.column_dimensions[chr(65 + idx)].width = max_length

            wb.save(file_name)

            # Step 5d: Display summary of results
            print(f"\nAnalysis for course: {selected_course}")
            print(f"Found {len(course_data)} sections")
            print(f"Results exported to {file_name}")
            break

        except (FileNotFoundError, ValueError, KeyError):
            print("Error processing course data")
            print(traceback.format_exc())
            continue
